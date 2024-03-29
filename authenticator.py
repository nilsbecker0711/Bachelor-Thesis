import librosa 
import numpy as np
import matplotlib.pyplot as plt
from sklearn.svm import SVC
from sklearn.model_selection import train_test_split
from sklearn.model_selection import GridSearchCV
import pandas as pd
import joblib
import os
from datetime import datetime
import openpyxl
from collections import Counter

dirname = os.path.dirname(__file__)


def extract_mfcc(audio_path):
    '''
    This method extracts the mel frequency cepstral coefficients (mfcc) from a given path to an audio file
    :param audio_path: path to audio file in .wav format
    :return: nbormalized frequency cepstral coefficients, sampling rate and audiodata
    '''
    audio, sr = librosa.load(audio_path)
    mfcc = librosa.feature.mfcc(y=audio, sr=sr, n_mfcc=128)

    mfcc_mean = np.mean(mfcc, axis=1)

    mfcc_mean_normalized = (mfcc_mean - np.min(mfcc_mean)) / (np.max(mfcc_mean) - np.min(mfcc_mean))
    #print(mfcc_mean_normalized)

    return mfcc_mean_normalized, sr, audio

def plot_mel_spectrogram(audio_path):
    
    '''
    Shows a visual representation of Mel Frequencys, by first calculating and then plotting them
    :param: audio_path path to audio file
    '''
   
    mfcc_mean_normalized, sr, audio = extract_mfcc(audio_path)


    mel_spec = librosa.feature.melspectrogram(y=audio, sr=sr, S=None, n_fft=2048, hop_length=512, n_mels=128, fmax=sr/2)
    mel_spec_db = librosa.power_to_db(mel_spec, ref=np.max)

    #Plotting
    plt.figure(figsize=(10, 4))
    librosa.display.specshow(mel_spec_db, sr=sr, x_axis='time', y_axis='mel')
    plt.colorbar(format='%+2.0f dB')
    plt.title('Mel Spectrogram')
    plt.tight_layout()
    plt.show()

def save_classifier(classifier, speaker_id, date):
    '''
    This method saves a given classifier to the models subfolder with current datetime as destinction method. The datetime format is D/M/Y_H_M
    :param classifier: classifier to save (As it is only SVC)
    :return: path to file
    '''

    try:
       filename = os.path.join(dirname, f'models\\svc_model{date}\\{speaker_id}.jl')
       joblib.dump(classifier, filename )
       return filename
    except Exception as e:
         print(e)
         return None
    
def load_classifiers(path, from_models=False):
    '''
   Load a pre-trained classifier
   :param path: path to the classifier
   :param from_models: if True the model to be loaded is from the models subfolder allowing a shorter path to be given
   :return: the pre-trained classifier or None if an Exception is raised 
    '''

    if from_models:
        path = os.path.join(dirname,f"models/{path}")   
    try:
        classifier=joblib.load(path)
        return classifier
    except Exception as e:
        print(e)
        return None

def train_speaker_classification(data_path, audio_path, tune=False, save = True):
    '''
    Trains a SVC to differentiate between speakers.
    :param data_path: link to an excel file containing paths to audio files and distict speaker ID's
    :param audio_path: path to folder, where audio files are stored
    :param tune: Indicates if the SVC hyperparameters C and kernel should be optimized before training
    :param save: Indicates if the model is to be saved
    :return: trained gender classifier
    '''
    data = os.path.join(dirname, data_path)
    voice_samples = pd.read_excel(data, usecols=[0,1])
    date = None
    if save:
        svc_paths =[]
        now = datetime.now()
        date = now.strftime("%d_%m_%Y_%H_%M")
        os.mkdir(os.path.join(dirname, f'models\\svc_model{date}'))

    speakers = []
    features = []
    audio_paths = []
    
    print(f"Feature Extraction started: {datetime.now()}")
    for index, sample in voice_samples.iterrows():
        if index == 1500:
            pass
        speaker_ID, audio_path_detail = sample
        audio_paths.append(audio_path+audio_path_detail)
        mfcc_features, sr, path = extract_mfcc(audio_path+audio_path_detail)
        speakers.append(speaker_ID)
        features.append(mfcc_features)  
    print(f"Feature Extraction completed: {datetime.now()}")
    
    ''' Filter out speakers with n<5 samples
    values = Counter(speakers)
    undo = []
    for speaker_number, amount in values.items(): 
        if amount < 5:
            for index, value in enumerate(speakers):
                if value == f'{speaker_number}':
                    undo.append(index)
    undo = sorted(undo, reverse=True)
    for delete in undo:
        speakers.pop(delete)
        features.pop(delete)
   '''

    values = Counter(speakers)
    classifiers = [] # holds all n classifiers for n distinct speakers
    amount_counter = 0
    counter = 0
    for speaker_number, amount in values.items():
        current_speaker = [0 for i in range(len(speakers))] #fill with zeros
        for i in range(amount_counter, amount_counter+amount):
            current_speaker[i] = 1
        amount_counter += amount
        runner = True
        runner_counter = 0
        while runner:
            try: #find invalid speaker data
                classifier = train_classifier(features, current_speaker, tune)
                runner = False
            except Exception as e:
                print(speaker_number, runner_counter)
                runner_counter += 1
                print(e)
                runner = True
        classifiers.append(classifier)
        
        if save:
            
            if counter < 10:
                svc_path = f'00{counter}-{speaker_number}'
                svc_paths.append(svc_path)
            elif counter < 100:
                svc_path = f'0{counter}-{speaker_number}'
                svc_paths.append(svc_path)
            else:
                svc_path = f'{counter}-{speaker_number}'
                svc_paths.append(svc_path)
            save_classifier(classifier, svc_path, date)
            counter += 1
    if save:

        build_svc_excel(date = date, speakers=speakers, audio_paths=audio_paths, svc_paths=svc_paths)
        return classifiers, speakers, svc_paths, audio_paths
    else:
        return classifiers

def train_clone_classifier(clone_path, audio_path, data_path):
    
    data = os.path.join(dirname, data_path)
    voice_samples = pd.read_excel(data, usecols=[0,1])
    speakers = []
    features = []
    print(f"Feature Extraction 1 started: {datetime.now()}")
    for index, sample in voice_samples.iterrows():
        _,audio_path_detail = sample
        mfcc_features, sr, path = extract_mfcc(audio_path+audio_path_detail)
        speakers.append(0)
        features.append(mfcc_features)
    print(f"Feature Extraction 1 completed: {datetime.now()}")
    print(f"Feature Extraction 2 started: {datetime.now()}")
    for filename in os.listdir(clone_path):
          mfcc_features, sr, path = extract_mfcc(f'samples/cloned/{filename}')
          speakers.append(1)
          features.append(mfcc_features)
    print(f"Feature Extraction 2 completed: {datetime.now()}")
    classifier = train_classifier(features, speakers, True)
    now = datetime.now()
    date = now.strftime("%d_%m_%Y_%H_%M")
    os.mkdir(os.path.join(dirname, f'models\\svc_model{date}'))
    save_classifier(classifier, "clone", date)

    return classifier

def add_classifier(audio_path, speaker_name, tune, date): 
        features = []
        speaker_features =[]
        audio_paths = []
        for filename in os.listdir(audio_path):
            audio_paths.append(audio_path+filename)
            feature,_,_ = extract_mfcc(audio_path+filename)
            features.append(0)
            speaker_features.append(feature)
        
        data = os.path.join(dirname, f"models/svc_model{date}/data.xlsx")
        workbook = openpyxl.load_workbook(data)
        worksheet = workbook.active
        speaker_number = int(worksheet.cell(row=worksheet.max_row, column=3).value[:3]) + 1
        voice_samples = pd.read_excel(data, usecols=[0,1])
        for index, sample in voice_samples.iterrows():
            speaker_id, audio_path_detail = sample
            mfcc_features, sr, path = extract_mfcc(audio_path_detail)
            speaker_features.append(0)
            features.append(mfcc_features)

        runner = True
        classifier = None
        svc_path = None
        while runner:
            try:
                classifier = train_classifier(features, speaker_features, tune) 
                svc_path = f'{speaker_number}-{speaker_name}'
                save_classifier(classifier, svc_path, date)
                runner = False
            except:
                runner = True
       
        #update excel
        workbook = openpyxl.load_workbook(os.path.join(dirname, f"models/svc_model{date}/data.xlsx"))
        worksheet = workbook.active
        for audio_path in audio_paths:
            data = [speaker_name, f'samples/{audio_path[8:]}', f'{svc_path}']
            worksheet.append(data)
        workbook.save(os.path.join(dirname, f"models/svc_model{date}/data.xlsx"))
        workbook.close()
        return classifier

def build_svc_excel(date, speakers, audio_paths, svc_paths):
    
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    data = []
    svc_index = 0
    for i in range(len(speakers)):
        if (i > 1) and (speakers[i] != speakers[i-1]):
            svc_index += 1
        data.append([speakers[i], audio_paths[i], svc_paths[svc_index]])
    for row_data in data:
        worksheet.append(row_data)
    workbook.save(os.path.join(dirname, f"models/svc_model{date}/data.xlsx"))
    workbook.close()

def tune_SVC_hyperparameters(X_train, y_train):
    '''
    Find the optimal parameters for C and Kernel for a 
    '''
    classifier =  SVC(probability=True)
   
    param_grid = {
    'C': [0.1, 1, 10],
    'kernel':['poly', 'linear', 'rbf'],
    }
    print(f"Parameter Search Started (Using all CPU Cores in parallel): {datetime.now()}")
    grid_search = GridSearchCV(classifier, param_grid, cv=5, n_jobs=-1)
    grid_search.fit(X_train, y_train)
    print(f"Parameter Search Completed: {datetime.now()}")
    params = grid_search.best_params_
    print(f'Beste Parameter: {params}')
    return [params['C'], params['kernel']]


def train_classifier(features, criteria, tune=False):

    X_train, X_test, y_train, y_test = train_test_split(features, criteria, test_size=0.2)

    if tune:
        C, kernel = tune_SVC_hyperparameters(X_train, y_train)
        classifier = SVC(C=C, kernel=kernel)
    else:
        classifier = SVC(C=0.1, kernel='poly')

    classifier.fit(X_train, y_train)
    accuracy = classifier.score(X_test, y_test)
    print("Classifier Accuracy: {:.2f}%".format(accuracy * 100))

    return classifier

def predict_single_speaker(classifiers, audio_path, clone_classifier, proba=False):
    '''
    Predict for a single sample who the speaker is
    :param classifier: a pretrained classifier 
    :param audio_path: path to the audio file to be classified
    :param proba = indicates if the input SVCs are trained for probabilities
    :return: If no exception is raised -> A String with the name of the sample and the predicted speaker, True
             If an exception is raised -> A String with a short notice of the failed prediction, False
    '''
    new_mfcc_features,_,_ = extract_mfcc(audio_path)
    #test if the voice is cloned
    if ((bool(clone_classifier.predict([new_mfcc_features])[0]))):
        return([-1, -1],False)
    predictions = []
    indizes = []
    index = 0
    try:
        for classifier in classifiers:
            if proba:
                pred = classifier.predict_proba([new_mfcc_features])[0]
                print(pred, index)
            else:
                pred =  classifier.predict([new_mfcc_features])[0]
                if pred == 1:
                    indizes.append([index, classifier.decision_function([new_mfcc_features])[0]])
               
            predictions.append(pred)
            index += 1

        if (len(indizes) > 1):
            indizes.sort(key = lambda x: x[1], reverse=True)
            return (indizes, True)
        elif(len(indizes) == 1):
            return (indizes, True)
        else: return ([-1, -1], True)
    except Exception as e:
        print(e)
        return("Error",False)

